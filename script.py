import openpyxl
import pandas as pd

# open data from instagram
following_path = "following.xlsx"
followers_path = "followers.xlsx"
# create two workbook objects
following_wk_obj = openpyxl.load_workbook(following_path)
followers_wk_obj = openpyxl.load_workbook(followers_path)

# create two active sheet objects
following_sheet_obj = following_wk_obj.active
followers_sheet_obj = followers_wk_obj.active

# get value of maximum rows and columns in following sheet
following_row = following_sheet_obj.max_row
following_column = following_sheet_obj.max_column

# get value of maximum rows and columns in followers sheet
followers_row = followers_sheet_obj.max_row
followers_column = followers_sheet_obj.max_column


# function to convert ig data to be used in code
def convert_ig_excel_to_readable_data(sheet_obj_row, sheet_obj):
    # add data to a dictionary
    ig_dict = {}
    for index in range(1, sheet_obj_row + 1):
        if index % 2 != 0:
            cell_obj = sheet_obj.cell(row=index, column=1)
            ig_dict[cell_obj.value] = cell_obj.value
    return ig_dict


# add followers to a dict for comparison
followers_dict = convert_ig_excel_to_readable_data(followers_row, followers_sheet_obj)
following_dict = convert_ig_excel_to_readable_data(following_row, following_sheet_obj)


# check who you follow but they dont follow you back
count_not_followed_back = 0
for following in following_dict:
    if following_dict.get(following) not in followers_dict:
        count_not_followed_back += 1
        print("you are following " + following_dict.get(following) + " but they dont follow you back")
print("You follow " + str(len(following_dict)) + " and " + str(count_not_followed_back) +
      " dont follow you back")





